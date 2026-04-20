"""리포트 생성 모듈 — 분석 결과를 엑셀 리포트로 출력"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

from .utils import OUTPUT_DIR, MASTER_FILE, format_currency, ensure_dirs
from .analyzer import (
    monthly_summary, trend_analysis, category_analysis,
    detect_anomalies, cash_flow_forecast, generate_advice,
)


# 스타일 상수
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="맑은 고딕", bold=True, size=11, color="2F5496")
NORMAL_FONT = Font(name="맑은 고딕", size=10)
WARN_FONT = Font(name="맑은 고딕", size=10, color="CC0000", bold=True)
MONEY_FORMAT = '#,##0'
PERCENT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def _apply_header_style(ws, row, max_col):
    """헤더 행에 스타일 적용"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _apply_data_style(ws, start_row, end_row, max_col, money_cols=None):
    """데이터 영역 스타일 적용"""
    money_cols = money_cols or []
    for row in range(start_row, end_row + 1):
        fill = ALT_FILL if (row - start_row) % 2 == 1 else None
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if col in money_cols:
                cell.number_format = MONEY_FORMAT


def _write_summary_sheet(wb, summary, forecast, advice_list):
    """요약 시트 작성"""
    ws = wb.active
    ws.title = "요약"
    ws.sheet_properties.tabColor = "2F5496"

    # 제목
    ws.merge_cells("A1:F1")
    ws["A1"] = f"📊 월간 재무 요약 — {summary.get('월', '')}"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # 핵심 지표
    row = 3
    ws.cell(row=row, column=1, value="항목").font = SUBTITLE_FONT
    ws.cell(row=row, column=2, value="금액").font = SUBTITLE_FONT
    ws.cell(row=row, column=3, value="비율").font = SUBTITLE_FONT
    _apply_header_style(ws, row, 3)

    metrics = [
        ("총수입", summary.get("총수입", 0), ""),
        ("총지출", summary.get("총지출", 0), ""),
        ("순이익", summary.get("순이익", 0), ""),
        ("", "", ""),
        ("고정비", summary.get("고정비", 0), f"{summary.get('고정비비율', 0):.1f}%"),
        ("변동비", summary.get("변동비", 0), f"{summary.get('변동비비율', 0):.1f}%"),
        ("미분류 지출", summary.get("미분류지출", 0), ""),
        ("", "", ""),
        ("거래 건수", summary.get("거래건수", 0), ""),
    ]

    for i, (label, value, ratio) in enumerate(metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        cell_val = ws.cell(row=r, column=2, value=value)
        cell_val.font = NORMAL_FONT
        if isinstance(value, (int, float)) and value != 0:
            cell_val.number_format = MONEY_FORMAT
        ws.cell(row=r, column=3, value=ratio).font = NORMAL_FONT
        if label == "순이익" and isinstance(value, (int, float)) and value < 0:
            cell_val.font = WARN_FONT

    # 현금 흐름 예측
    row = row + len(metrics) + 2
    ws.cell(row=row, column=1, value="💰 현금 흐름 예측").font = SUBTITLE_FONT
    row += 1
    forecast_items = [
        ("기준기간", forecast.get("기준기간", "")),
        ("월평균 고정비", forecast.get("월평균고정비", 0)),
        ("월평균 변동비", forecast.get("월평균변동비", 0)),
        ("월평균 총지출", forecast.get("월평균총지출", 0)),
        ("월평균 수입", forecast.get("월평균수입", 0)),
        ("예상 순이익", forecast.get("예상순이익", 0)),
        ("최소 필요 자금", forecast.get("최소필요자금", 0)),
        ("권장 보유 자금 (20% 여유)", forecast.get("권장보유자금", 0)),
    ]
    for label, value in forecast_items:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = NORMAL_FONT
        if isinstance(value, (int, float)):
            cell.number_format = MONEY_FORMAT
        row += 1

    # 조언
    row += 1
    ws.cell(row=row, column=1, value="📋 재무 조언").font = SUBTITLE_FONT
    row += 1
    for adv in advice_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=adv).font = NORMAL_FONT
        ws.row_dimensions[row].height = 25
        row += 1

    # 지출 TOP 10
    row += 1
    ws.cell(row=row, column=1, value="📈 지출 항목 TOP 10").font = SUBTITLE_FONT
    row += 1
    headers = ["순위", "대분류", "중분류", "금액"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))

    for rank, item in enumerate(summary.get("지출TOP10", []), 1):
        row += 1
        ws.cell(row=row, column=1, value=rank).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=item["대분류"]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=item["중분류"]).font = NORMAL_FONT
        cell = ws.cell(row=row, column=4, value=item["금액"])
        cell.font = NORMAL_FONT
        cell.number_format = MONEY_FORMAT

    # 컬럼 너비
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


def _write_transactions_sheet(wb, df, month=None):
    """거래내역 시트 작성"""
    ws = wb.create_sheet("거래내역")
    ws.sheet_properties.tabColor = "4472C4"

    if month:
        df_show = df[df["연월"] == month].copy() if "연월" in df.columns else df.copy()
    else:
        df_show = df.copy()

    cols_to_show = ["거래일자", "거래유형", "적요", "거래처", "입금액", "출금액", "대분류", "중분류"]
    df_show = df_show[[c for c in cols_to_show if c in df_show.columns]]

    # 거래일자를 문자열로
    if "거래일자" in df_show.columns:
        df_show["거래일자"] = pd.to_datetime(df_show["거래일자"]).dt.strftime("%Y-%m-%d")

    # 헤더
    for col_idx, col_name in enumerate(df_show.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    _apply_header_style(ws, 1, len(df_show.columns))

    # 데이터
    for row_idx, row_data in enumerate(df_show.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    money_cols = []
    for i, c in enumerate(df_show.columns, 1):
        if c in ("입금액", "출금액"):
            money_cols.append(i)
    _apply_data_style(ws, 2, len(df_show) + 1, len(df_show.columns), money_cols)

    # 컬럼 너비
    widths = {"거래일자": 12, "거래유형": 8, "적요": 30, "거래처": 20,
              "입금액": 15, "출금액": 15, "대분류": 10, "중분류": 12}
    for i, c in enumerate(df_show.columns):
        ws.column_dimensions[chr(65 + i)].width = widths.get(c, 12)


def _write_cost_analysis_sheet(wb, cat_analysis, cost_type):
    """고정비/변동비 분석 시트"""
    sheet_name = f"{cost_type}분석"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "ED7D31" if cost_type == "변동비" else "70AD47"

    df_cat = cat_analysis.get(cost_type)
    if df_cat is None or df_cat.empty:
        ws.cell(row=1, column=1, value=f"{cost_type} 데이터가 없습니다.")
        return

    # 테이블
    headers = ["항목", "금액", "비율(%)", "전월금액", "증감액", "증감률(%)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))

    for row_idx, (cat_name, row_data) in enumerate(df_cat.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=cat_name).font = NORMAL_FONT
        ws.cell(row=row_idx, column=2, value=row_data["금액"]).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=3, value=round(row_data["비율"], 1))
        ws.cell(row=row_idx, column=4, value=row_data["전월금액"]).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=5, value=row_data["증감액"]).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=6, value=round(row_data["증감률"], 1))

    _apply_data_style(ws, 2, len(df_cat) + 1, len(headers), [2, 4, 5])

    # 파이 차트 (비율)
    if len(df_cat) >= 2:
        pie = PieChart()
        pie.title = f"{cost_type} 구성 비율"
        pie.style = 10
        pie.width = 16
        pie.height = 10

        labels = Reference(ws, min_col=1, min_row=2, max_row=len(df_cat) + 1)
        data = Reference(ws, min_col=2, min_row=1, max_row=len(df_cat) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        ws.add_chart(pie, f"H2")

    # 컬럼 너비
    for i, w in enumerate([15, 15, 10, 15, 15, 12], 0):
        ws.column_dimensions[chr(65 + i)].width = w


def _write_trend_sheet(wb, trend_df):
    """추세 분석 시트"""
    ws = wb.create_sheet("추세")
    ws.sheet_properties.tabColor = "FFC000"

    if trend_df is None or trend_df.empty:
        ws.cell(row=1, column=1, value="추세 데이터가 부족합니다.")
        return

    headers = ["월", "총수입", "총지출", "고정비", "변동비", "순이익"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_style(ws, 1, len(headers))

    for row_idx, (_, row_data) in enumerate(trend_df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row_data["연월"]).font = NORMAL_FONT
        for col_idx, col_name in enumerate(["총수입", "총지출", "고정비", "변동비", "순이익"], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, 0))
            cell.number_format = MONEY_FORMAT
            cell.font = NORMAL_FONT

    _apply_data_style(ws, 2, len(trend_df) + 1, len(headers), [2, 3, 4, 5, 6])

    # 라인 차트
    if len(trend_df) >= 2:
        chart = LineChart()
        chart.title = "월별 자금 흐름 추이"
        chart.style = 10
        chart.width = 24
        chart.height = 14
        chart.y_axis.title = "금액 (원)"
        chart.x_axis.title = "월"

        cats = Reference(ws, min_col=1, min_row=2, max_row=len(trend_df) + 1)
        for col_idx in range(2, 7):
            values = Reference(ws, min_col=col_idx, min_row=1, max_row=len(trend_df) + 1)
            chart.add_data(values, titles_from_data=True)

        chart.set_categories(cats)
        ws.add_chart(chart, "A" + str(len(trend_df) + 4))

    for i, w in enumerate([10, 15, 15, 15, 15, 15], 0):
        ws.column_dimensions[chr(65 + i)].width = w


def _write_alert_sheet(wb, anomalies, advice_list):
    """경고 및 조언 시트"""
    ws = wb.create_sheet("경고")
    ws.sheet_properties.tabColor = "FF0000"

    ws.cell(row=1, column=1, value="🔔 이상 항목 탐지").font = TITLE_FONT
    ws.merge_cells("A1:F1")

    if anomalies:
        headers = ["대분류", "중분류", "이번달 금액", "평균 금액", "배율", "초과액"]
        row = 3
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        _apply_header_style(ws, row, len(headers))

        for a in anomalies:
            row += 1
            ws.cell(row=row, column=1, value=a["대분류"]).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=a["중분류"]).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=a["이번달"]).number_format = MONEY_FORMAT
            ws.cell(row=row, column=4, value=a["평균"]).number_format = MONEY_FORMAT
            ws.cell(row=row, column=5, value=round(a["배율"], 1)).font = WARN_FONT
            ws.cell(row=row, column=6, value=a["초과액"]).number_format = MONEY_FORMAT
    else:
        ws.cell(row=3, column=1, value="이상 항목이 없습니다.").font = NORMAL_FONT

    row = (3 + len(anomalies) + 3) if anomalies else 5
    ws.cell(row=row, column=1, value="📋 종합 재무 조언").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for adv in advice_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=adv).font = NORMAL_FONT
        ws.row_dimensions[row].height = 28
        row += 1

    for i, w in enumerate([12, 12, 15, 15, 8, 15], 0):
        ws.column_dimensions[chr(65 + i)].width = w


def generate_report(month=None):
    """
    월간 분석 리포트 생성.
    반환: 생성된 파일 경로
    """
    ensure_dirs()

    if not Path(MASTER_FILE).exists():
        raise FileNotFoundError("마스터 데이터가 없습니다. 먼저 데이터를 가져오세요.")

    df = pd.read_excel(MASTER_FILE)
    df["거래일자"] = pd.to_datetime(df["거래일자"], errors="coerce")
    df["연월"] = df["거래일자"].dt.to_period("M").astype(str)

    if month is None:
        month = df["연월"].max()

    # 분석 실행
    summary = monthly_summary(df, month)
    trend = trend_analysis(df)
    cat_analysis = category_analysis(df, month)
    anomalies = detect_anomalies(df, month)
    forecast = cash_flow_forecast(df)
    advice = generate_advice(summary, anomalies, forecast)

    # 워크북 생성
    wb = Workbook()

    _write_summary_sheet(wb, summary, forecast, advice)
    _write_transactions_sheet(wb, df, month)
    _write_cost_analysis_sheet(wb, cat_analysis, "고정비")
    _write_cost_analysis_sheet(wb, cat_analysis, "변동비")
    _write_trend_sheet(wb, trend)
    _write_alert_sheet(wb, anomalies, advice)

    # 저장
    output_path = OUTPUT_DIR / f"리포트_{month}.xlsx"
    wb.save(output_path)

    return str(output_path)
